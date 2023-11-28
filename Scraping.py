import time
import requests
import asyncio
import sys
import json
import datetime
import gspread

from datetime import date
from openpyxl.styles import Alignment
from googleapiclient.discovery import build  # Added
from google.oauth2 import service_account

from selenium import webdriver
from gspread_formatting import batch_updater

month = sys.argv[1]
day = sys.argv[2]

# Get Data from purpose site
async def scanData(start_month, start_day):
    # Send get request
    async def send_request(p_url):
        url = p_url
        headers = {
            'Content-Type': 'application/json',
            "Accept": "application/json",
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.82 Safari/537.36',
        }
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            json_response = response.text
            return json_response
        else:
            return ''

    # Append to array
    async def append_to_arr(store, data):
        length = len(store)
        if(length == 0):
            length = 1
        for i in range(len(data)):
            store.append([length + i, data[i]['userInfo']['displayName'], data[i]['score']])
        return store

    # Get Ranking List
    async def getRankingList(containerID):
        result = []
        url = f'https://api-dsa.17app.co/api/v1/leaderboards/eventory?containerID={containerID}&cursor=&count=100'
        nextCursor = 'init'
        while nextCursor != '':
            ranking_list = await send_request(url)
            ranking_list = json.loads(ranking_list)
            if 'data' in ranking_list:
                await append_to_arr(result, ranking_list['data'])
            nextCursor = ranking_list['nextCursor']
            url = f'https://api-dsa.17app.co/api/v1/leaderboards/eventory?containerID={containerID}&cursor={nextCursor}&count=100'

        return result

    # Create New Google Sheet
    async def createGoogleSheet(filename):
        SCOPES = ['https://www.googleapis.com/auth/drive']  # Modified
        credentials = service_account.Credentials.from_service_account_file('service-account.json', scopes=SCOPES)

        drive = build('drive', 'v3', credentials=credentials)
        file_metadata = {
            'name': filename,
            'parents': '11seQXAOIxXozPsCy7rG_CgJW0L8rdPmM',
            # 'parents': ['1EO1rQPYbTRGUQl4mGvkFAY2cm0zn947w'],
            'mimeType': 'application/vnd.google-apps.spreadsheet'
        }
        res = drive.files().create(body=file_metadata).execute()
        permission_body = {
            'role': 'writer',  # Set the desired role ('reader', 'writer', 'commenter', 'owner')
            'type': 'anyone',  # Share with anyone
        }
        drive.permissions().create(fileId=res['id'], body=permission_body).execute()

        file_id = res['id']
        folder_id = '11seQXAOIxXozPsCy7rG_CgJW0L8rdPmM'

        # Update the file's parents
        # drive.files().update(fileId=file_id, addParents=folder_id).execute()
        return file_id

    # Get Sheet ID by file name
    async def get_sheet_by_name(file_name, folder_name):
        SCOPES = ['https://www.googleapis.com/auth/drive']  # Modified
        credentials = service_account.Credentials.from_service_account_file('service-account.json', scopes=SCOPES)
        drive_service = build('drive', 'v3', credentials=credentials)
        sheet_id = None

        # results = drive_service.files().list(q="name='" + file_name + "' and mimeType='application/vnd.google-apps.spreadsheet' and '" + folder_name + "' in parents",
        #                                     pageSize=10, fields="nextPageToken, files(id, name)").execute()
        results = drive_service.files().list(q="name='" + file_name + "' and mimeType='application/vnd.google-apps.spreadsheet' ",
                                pageSize=10, fields="nextPageToken, files(id, name)").execute()
        items = results.get('files', [])
        if not items:
            return ''
        else:
            sheet_id = items[0]['id']

        return sheet_id

    # Get Calculate result
    def calculate_date(start_month, start_day):
        current_year = datetime.datetime.now().year
        current_month = datetime.datetime.now().month
        current_day = datetime.datetime.now().day

        date1 = datetime.datetime.strptime(f'{current_year}-{current_month}-{current_day}', '%Y-%m-%d')
        if(int(month) > month):
            start_year = current_year - 1
        else:
            start_year = current_year
        date2 = datetime.datetime.strptime(f'{start_year}-{start_month}-{start_day}', '%Y-%m-%d')
        delta = date1 - date2
        return delta.days
    
    # Create new sheet into spreadsheet
    async def create_sheet_into_spreadsheet(sheetID, data):
        SCOPES = ['https://www.googleapis.com/auth/drive']
        SERVICE_ACCOUNT_FILE = 'service-account.json'

        creds = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(sheetID)

        data = data['Data']
        for i in range(len(data)):
            worksheet = spreadsheet.add_worksheet(title=data[i]['EventID'], rows='100', cols='100')

    # Insert image into worksheet
    async def insert_image_in_googlesheet(sheetID, image):
        SCOPES = ['https://www.googleapis.com/auth/drive']
        SERVICE_ACCOUNT_FILE = 'service-account.json'

        creds = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(sheetID)

        sheet = spreadsheet.sheet1
        sheet.update_title('タイトル')

        sheet_range = "A1"
        batch = batch_updater(sheet.spreadsheet)
        batch.set_row_height(sheet, '1:1', 440)
        batch.set_column_width(sheet, 'A:A', 500)
        batch.execute()
        
        insert_image = f"=IMAGE(\"{image}\", 1)"
        sheet.update(sheet_range, [[insert_image]], value_input_option="USER_ENTERED")

    # Insert html content into worksheet
    async def insert_content_in_googlesheet(sheetID, element, parent_title, title):
        # result = None
        # content = element.find_elements('css selector', '.btCdvi')
        # if(len(content) > 0):
        #     result = content[0]
        # else:
        #     content = element.find_elements('css selector', '.bjzlAe')
        #     result = [el.get_attribute('innerHTML') for el in content]

        # worksheet = spreadsheet.add_worksheet(title=f"{parent_title} - {title}", rows='100', cols='100')
        # sheet_range = "A1"
        # worksheet.update(sheet_range, result, value_input_option="USER_ENTERED")
        search_panel = element.find_elements('css selector', '.bpEaZC')
        if(len(search_panel) > 0):
            return 
        inner_html = ''
        content = element.find_elements('css selector', '.bjzlAe')
        if(len(content)):
            for i in range(len(content)):
                inner_html += content[i].get_attribute('outerHTML')
        print("************************************************")
        print(inner_html)
        print("************************************************")

    # Get attr of element
    async def handleGetAttr(elements, type):
        for element in elements:
            url = element.get_attribute(type)
            return url

    # Insert jpg, html content into spreadsheet
    async def insert_image(sheetID, event_id):
        chrome_options = webdriver.ChromeOptions()
        browser = webdriver.Chrome(options=chrome_options)

        # Maximize the browser window
        browser.maximize_window()

        # Bring the browser window to the front
        browser.execute_script("window.focus();")
        time.sleep(3)
        
        browser.get(f'https://event.17.live/{event_id}')
        time.sleep(10)
        main_image_elements = browser.find_elements('css selector', '.sc-crHlIS')
        main_image = await handleGetAttr(main_image_elements, 'src')

        await insert_image_in_googlesheet(sheetID, main_image)

        tab_elements = browser.find_elements('css selector', '.kGvAFP')
        for i in range(len(tab_elements)):
            if(i == 0):
                continue
            tab_title = tab_elements[i].text
            print(tab_title)
            print('================')
            tab_elements[i].click()

            sub_tab_group = browser.find_elements('css selector', '.gOMukq')
            if(len(sub_tab_group) > 0):
                sub_tab_elements = sub_tab_group[0].find_elements('css selector', '.ffjCOc')
                if(len(sub_tab_elements) > 0):
                    for j in range(len(sub_tab_elements)):
                        sub_tab_title = sub_tab_elements[j].text
                        sub_tab_elements[j].click()
                        print(f"{sub_tab_title}")
                        # 
                        last_sub_tab_group = browser.find_elements('css selector', '.gOMukq')
                        if(len(last_sub_tab_group) > 1):
                            last_sub_tab_elements = last_sub_tab_group[1].find_elements('css selector', '.ffjCOc')
                            if(len(last_sub_tab_elements) > 0):
                                print('---------------------------------')
                                for l in range(len(last_sub_tab_elements)):
                                    sub_tab_title = last_sub_tab_elements[l].text
                                    last_sub_tab_elements[l].click()
                                    print(f"{sub_tab_title} - {l}")
                                    await insert_content_in_googlesheet(sheetID, browser, tab_title, sub_tab_title)
                                print('---------------------------------')
                        # 
                        await insert_content_in_googlesheet(sheetID, browser, tab_title, sub_tab_title)
                    print('------------()()()()()----------')

    def write_into_googlesheet(sheetID, data):
        SCOPES = ['https://www.googleapis.com/auth/drive']
        SERVICE_ACCOUNT_FILE = 'service-account.json'

        creds = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(sheetID)

        data = data['Data']
        for i in range(len(data)):
            worksheet = spreadsheet.worksheet(data[i]['EventID'])
            start_row = 1  # Replace with the starting row index
            start_col = 2  # Replace with the starting column index
            end_row = 1    # Replace with the ending row index
            end_col = 3    # Replace with the ending column index

            # Merge the cells
            worksheet.merge_cells(start_row, start_col, end_row, end_col)

            # Write text in the merged cell
            merged_cell = worksheet.cell(start_row, start_col)  # Use the top-left cell of the merged range

            current_month = datetime.datetime.now().month
            current_day = datetime.datetime.now().day

            merged_cell.value = f"{current_month}月{current_day}日"
            merged_cell.alignment = Alignment(horizontal='center')

            # Update the sheet with the modified cell
            worksheet.update_cells([merged_cell])

            row_index = 2  # Assuming you want to insert the data in the 2nd row

            worksheet.insert_rows(data[i]['List'], row=row_index)

    # Get event all url
    url = 'https://wap-api.17app.co/api/v1/event?region=JP&status=1'
    
    event_urls = None
    try:
        event_urls = await send_request(url)
    except Exception as e:
        print(f"An error occurred while fetching JSON for {url}: {e}")

    event_urls = json.loads(event_urls)
    event_urls = event_urls['events']['inProgress']

    event_url_arr = []
    if (len(event_urls) > 1):
        for i in range(len(event_urls)):
            event_url_arr.append(event_urls[i]['descriptionURL'])

    # Get containerID from event refrence api
    event_json_data = []
    for i in range(len(event_url_arr)):
        event_id = event_url_arr[i].split('/')[-1]
        json_url = f'https://webcdn.17app.co/campaign/projects/{event_id}/references.json'
        try:
            event_json_response = await send_request(json_url)
            if(event_json_response != ''):
                data = json.loads(event_json_response)
                data = data['fetcher']
                # Get Current Date
                current_date = date.today()
                formatted_date = current_date.strftime("%Y-%m-%d")

                # Get event data.
                event_data = []
                for i in range(len(data)):
                    event_data.append({
                        "EventID": data[i]['id'][12:],
                        "ContainerID": data[i]['value']['args'][0],
                        "List": []
                    })

                res = {
                    "ID": event_id,
                    "Date": formatted_date,
                    "Data": event_data,
                    "Count": 0
                }
                event_json_data.append(res)
        except Exception as e:
            print(f"An error occurred while fetching JSON for {json_url}: {e}")

    for i in range(len(event_json_data)):
        data = event_json_data[i]['Data']
        for j in range(len(data)):
            event_json_data[i]['Data'][j]['List'] = await getRankingList(data[j]['ContainerID'])

    for i in range(len(event_json_data)):
        current_year = datetime.datetime.now().year
        current_month = datetime.datetime.now().month
        current_day = datetime.datetime.now().day

        filename = f"Ranking_{event_json_data[i]['ID']}_{current_year}_{start_month}_{start_day}"

        if(int(start_month) == current_month and int(start_day) == current_day):
            # create new sheet
            sheetID = await createGoogleSheet(filename)
            print(sheetID)
            event_json_data[i]['Count'] == 0
            await insert_image(sheetID, event_json_data[i]['ID'])
            await create_sheet_into_spreadsheet(sheetID, event_json_data[i])
            write_into_googlesheet(sheetID, event_json_data[i])
        else:
            folder_name = '11seQXAOIxXozPsCy7rG_CgJW0L8rdPmM'
            sheetID = await get_sheet_by_name(filename, folder_name)
            print(sheetID)
            event_json_data[i]['Count'] = calculate_date(start_month, start_day)
            write_into_googlesheet(sheetID, event_json_data[i])

    return event_json_data

# Create New Google Sheet
async def createGoogleSheet(filename):
    SCOPES = ['https://www.googleapis.com/auth/drive']  # Modified
    credentials = service_account.Credentials.from_service_account_file('service-account.json', scopes=SCOPES)

    drive = build('drive', 'v3', credentials=credentials)
    file_metadata = {
        'name': filename,
        'parents': '11seQXAOIxXozPsCy7rG_CgJW0L8rdPmM',
        # 'parents': ['1EO1rQPYbTRGUQl4mGvkFAY2cm0zn947w'],
        'mimeType': 'application/vnd.google-apps.spreadsheet'
    }
    res = drive.files().create(body=file_metadata).execute()
    permission_body = {
        'role': 'writer',  # Set the desired role ('reader', 'writer', 'commenter', 'owner')
        'type': 'anyone',  # Share with anyone
    }
    drive.permissions().create(fileId=res['id'], body=permission_body).execute()

    file_id = res['id']
    folder_id = '11seQXAOIxXozPsCy7rG_CgJW0L8rdPmM'

    # Update the file's parents
    # drive.files().update(fileId=file_id, addParents=folder_id).execute()
    return file_id

async def main():
    # result = await scanData(month, day)
    result = await createGoogleSheet('test')
    return result

asyncio.run(main())
